from django.shortcuts import render
from ..permisions import has_permision
from ..models import Part, Order, BOM
from pandas import read_excel
from django.contrib.auth.decorators import login_required
from json import loads
from openpyxl.utils import get_column_letter
from django.utils import timezone
from datetime import timezone
from django.db import transaction
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from collections import defaultdict




@login_required
def part_add(request):
    has_permision(request, 'admin')
    orders = Order.objects.all()
    if request.method == 'POST':
        new_order = request.POST.get('new_order')
        old_order = request.POST.get('old_order')
        uploaded_file = request.FILES.get('file')
        try:
            with transaction.atomic():
                df = read_excel(uploaded_file)
                if new_order:
                    if Order.objects.filter(order_number=new_order).exists():
                        return render(request, 'upload_part.html', {"error":f'{new_order} order already exist', 'orders':orders})
                    order = Order.objects.create(order_number=new_order)
                elif old_order:
                    order = Order.objects.get(order_number=old_order)
                else:
                    return render(request, 'upload_part.html', {"error":'please choose an order', 'orders':orders})
                all_part = []
                for _, part in df.iterrows():
                    description=part['description']
                    if not description:
                        description = '-'
                    id=part['anbar_code']
                    my_part = Part.objects.create(id=id, part_number=str(part['partnumber']).strip(),count=int(part['quantity']),description=description, order=order)
                    all_part.append(my_part)
        except Exception as e:
            return render(request, 'upload_part.html', {"error":str(e), 'orders':orders})
        return render(request, 'part_postview.html', {'parts': all_part})
    #GET
    return render(request, 'upload_part.html', {'orders':orders})


@login_required
def download_part_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts Template"
    ws.append(["partnumber", "quantity", "description","anbar_code"])
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 15
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=part_template.xlsx'
    wb.save(response)
    return response


@login_required
def part_list(request):
    has_permision(request, 'anbardar')

    if request.method == "POST":
        try:
            data = loads(request.body)
            code_anbar = data.get('code_anbar')
            failure = data.get('failure')
            description = data.get('description')
            count = data.get('count')
            failure_mapping = {'accepted':'ACCEPT','rejected':'REJECT','conditional type_1':'COND1','conditional type_2':'COND2','conditional type_3':'COND3'}
            
            with transaction.atomic():
                part = Part.objects.get(pk=code_anbar)
                part.failure = failure_mapping[failure]
                part.description = description
                if count:
                    part.count = count
                    part.changed_request = 'manual'
                part.save()

            return JsonResponse({'status': 'success'})
        except Part.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Part not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    if has_permision(request, 'admin')==True:
        parts = Part.objects.all()
    else:
        parts = Part.objects.all()#filter(pk__lt=10101021016)######################################################################filter tolidi

    order_number = request.GET.get('order_number')
    part_number = request.GET.get('part_number')
    code_anbar = request.GET.get('code_anbar')
    failure = request.GET.get('failure')
    page_number = request.GET.get('page', 1)

    if order_number:
        parts = parts.filter(order__order_number__icontains=order_number)
    if part_number:
        parts = parts.filter(part_number__icontains=part_number)
    if code_anbar:
        parts = parts.filter(pk__icontains=code_anbar)
    if failure:
        parts = parts.filter(failure__in=failure.split(','))

    paginator = Paginator(parts, 20)
    page_obj = paginator.get_page(page_number)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        part_data = [
            {
                'code_anbar': p.pk,
                'order_number': p.order.order_number,
                'part_number': p.part_number,
                'failure' : p.get_failure_display(),
                'description' : p.description,
                'count' : p.count,
                'reserve' : p.reserve,
                'ordered' : p.freeze,
            }
            for p in page_obj
        ]
        return JsonResponse({
            'parts': part_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })

    return render(request, 'part_list.html', {
        'parts': page_obj,
    })



@login_required
def export_excel_logs(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'logs'
    headers = ['Code Anbar', 'Part Number', 'Timestamp', 'Request ID', 'Changed By', 'Changes']
    ws.append(headers)
    parts = Part.objects.all()
    for part in parts:
        changes = part.get_history_changes()
        if changes:
            for c in changes[0]["changes"].changes:
                history_date = changes[0]['history_instance'].history_date
                history_date_naive = history_date.astimezone(timezone.utc).replace(tzinfo=None)
                ws.append([part.id,part.part_number,history_date_naive,part.changed_request,changes[0]['history_instance'].history_user.username if changes[0]['history_instance'].history_user else '',f"{c.field} {c.old} → {c.new}"])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=part-logs.xlsx'
    wb.save(response)
    return response




@login_required
def part_history(request):
    has_permision(request, 'admin')
    code_anbar_query = request.GET.get('code_anbar', '').strip()
    request_id_query = request.GET.get('request_id', '').strip()
    all_parts = Part.objects.all()

    if code_anbar_query:
        all_parts = all_parts.filter(id__icontains=code_anbar_query)

    if request_id_query:
        all_parts = all_parts.filter(changed_request__icontains=request_id_query)

    object_changes = []
    for part in all_parts:
        changes = part.get_history_changes()
        if changes:
            object_changes.append({
                'object': part,
                'changes': changes
            })
    paginator = Paginator(object_changes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'history.html', {
        'page_obj': page_obj,
        'code_anbar_query': code_anbar_query,
        'request_id_query': request_id_query,
    })


@login_required
def part_bom(request):
    has_permision(request, 'anbardar')

    boms = BOM.objects.filter(board__status='NR')

    parts = Part.objects.filter(part_number__in=boms.values_list('part_number',flat=True)).distinct()
    order_number = request.GET.get('order_number')
    part_number = request.GET.get('part_number')
    code_anbar = request.GET.get('code_anbar')
    designator = request.GET.get('designator')
    name = request.GET.get('name')
    project = request.GET.get('project')
    page_number = request.GET.get('page', 1)

    if order_number:
        parts = parts.filter(order__order_number__icontains=order_number)
    if part_number:
        parts = parts.filter(part_number__icontains=part_number)
    if code_anbar:
        parts = parts.filter(pk__icontains=code_anbar)
    if designator:
        boms = boms.filter(designators__icontains=designator)
    if name:
        boms = boms.filter(board__title__icontains=name)
    if project:
        boms = boms.filter(board__project_name__name__icontains=project)   
    
    part_lookup = defaultdict(list)
    for part in parts:
        part_lookup[part.part_number].append(part)
    merged_data = defaultdict(list)
    for bom in boms:
        for part in part_lookup.get(bom.part_number, []):
            merged_data[bom.part_number].append({
                'project': bom.board.project_name.name,
                'board': f'E{bom.board.name}V{bom.board.version}({bom.board.title})',
                'designators': bom.designators,
                'part_number':bom.part_number,
                'code_anbar': part.id,
                'order_number': part.order.order_number,
                'failure': part.failure,
            })

    merged_data_list = [item for sublist in merged_data.values() for item in sublist]

    paginator = Paginator(merged_data_list, 20)
    page_obj = paginator.get_page(page_number)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        part_data = [item for item in page_obj]
        return JsonResponse({
            'parts': part_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })

    return render(request, 'part_bom_list.html', {
        'parts': page_obj,
    })