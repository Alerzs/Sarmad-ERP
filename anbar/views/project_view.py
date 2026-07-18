from django.shortcuts import render, redirect
from ..permisions import has_permision
from ..models import Project, BOM, Board
from pandas import read_excel
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.utils.safestring import mark_safe
from json import dumps
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@login_required
def flow(request):
    has_permision(request, 'admin')
    nodes = [
        {"id": "دریافت", "x": 100, "y": 80,  "url": "https://example.com/a"},
        {"id": "B", "x": 100, "y": 180, "url": "https://example.com/b"},
        {"id": "C", "x": 100, "y": 280, "url": "https://example.com/c"},
        {"id": "D", "x": 200, "y": 380, "url": "https://example.com/c"},
    ]

    links = [
        {"id": "دریافت-C", "source": "دریافت", "target": "C", "shape": "curve"},
        {"id": "B-C", "source": "B", "target": "C", "shape": "line"},
        {"id": "D-دریافت", "source": "D", "target": "دریافت", "shape": "curve"},
    ]
    by_id = {n["id"]: n for n in nodes}

    def add_control_point(link, offset=120):
        s = by_id[link["source"]]
        t = by_id[link["target"]]
        midx = (s["x"] + t["x"]) / 2
        midy = (s["y"] + t["y"]) / 2

        cx = midx + offset
        cy = midy
        link["cx"] = cx
        link["cy"] = cy

    for lk in links:
        if lk.get("shape") == "curve":
            add_control_point(lk, offset=120)

    return render(request, "flow_chart.html", {
        "nodes_json": dumps(nodes),
        "links_json": dumps(links),
    })



@login_required
def project_list(request):
    has_permision(request, 'admin')
    projects = Project.objects.all()
    return render(request, 'project_list.html', {'projects':projects})



@login_required
def project_add(request):
    has_permision(request, 'admin')
    
    if request.method == 'POST':

        projectname = request.POST.get('projectname')
        boardname = request.POST.get('boardname')
        board_id = request.POST.get('id')
        version = request.POST.get('version')
        uploaded_file = request.FILES.get('file')
        try:
            with transaction.atomic():
                df = read_excel(uploaded_file)
                project, created = Project.objects.get_or_create(name=projectname)
                my_board, board_created = Board.objects.get_or_create(title=boardname, name=board_id, version=version, project_name=project, status='NR')
                boms = []
                if not board_created:
                    BOM.objects.filter(board=my_board).delete()
                for _, bom in df.iterrows():
                    my_bom = BOM.objects.create(board=my_board ,part_number=bom['Part Number'],designators=bom['Designator'],description="-", count=bom['Quantity'])#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                    boms.append(my_bom)
        except Exception as e:
            return render(request, 'project_add.html', {"error":str(e)})
        return render(request, 'project_postview.html', {'boms':boms, 'data':{'projectname':projectname,'boardname':boardname,'id':board_id,'version':version}})
           
    board_qs = Board.objects.values_list('project_name__name','title', 'name').distinct()
    board_map = board_map = {f"{p}::{t}": bid for p, t, bid in board_qs}
    projects = Project.objects.all()

    return render(request, 'project_add.html', {'board_map_json': mark_safe(dumps(board_map)), "projects" : projects})



@login_required
def download_bom_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Template"
    ws.append(["Part Number", "Designator", "Quantity"])
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 15
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=bom_template.xlsx'
    wb.save(response)
    return response