from django.shortcuts import render, redirect
from ..permisions import has_permision
from ..models import Darkhast, PaletAnbar, Palet, Part, Board, BOM, User, Project, Order, Mafghood
from django.contrib.auth.decorators import login_required
from json import dumps
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum, F, Count
from collections import defaultdict
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.http import HttpResponse
from openpyxl import Workbook



@login_required
def export_excel_req(request, pk):
    has_permision(request, 'anbardar')
    my_req = get_object_or_404(Darkhast, id=pk)
    my_palets = PaletAnbar.objects.filter(palet__req=my_req).distinct()
    wb = Workbook()
    ws = wb.active
    ws.title = str(my_req.id)+ 'request'
    ws.merge_cells('A1:D1')
    ws['A1'] = f'request{my_req.pk}/{my_req.palet_set.first().bom.board.title}'

    headers = ['pallet number', 'code anbar', 'quantity', 'partnumber']
    ws.append(headers)
    for palet in my_palets:
        ws.append([palet.number, palet.palet_set.first().part.pk, palet.cal_qnt(), palet.palet_set.first().part.part_number])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={my_req.id}_request.xlsx'
    wb.save(response)
    return response

@login_required
def darkhast(request):
    if request.method == 'POST':
        selected_order = request.POST.getlist('order')
        selected_user = request.POST.get('user')
        board_ids = request.POST.getlist('board')
        boms = list(BOM.objects.filter(board__in=board_ids, status='EPT').select_related('board'))
        target_user = User.objects.get(id=selected_user)
        my_req = Darkhast.objects.create(user=request.user, for_user=target_user, req_type='ASS')

        parts_qs = (
            Part.objects
            .filter(order__order_number__in=selected_order)
            .exclude(failure='REJECT')
            .select_related('order')
            .order_by('pk')
        )
        parts_by_pn = defaultdict(list)
        for part in parts_qs:
            parts_by_pn[part.part_number].append(part)
        rejected_parts = {
            p.part_number: p.count
            for p in Part.objects.filter(failure='REJECT')
        }

        insf, rej = [], []
        palets_to_create = []
        parts_to_update = set()

        with transaction.atomic():
            for bom in boms:
                Pn = bom.part_number
                required_count = int(bom.count)
                available_parts = parts_by_pn.get(Pn, [])
                if bom.board.status == 'RC':
                    available_count = sum(p.count - p.reserve for p in available_parts)
                else:
                    available_count = sum(p.count - p.reserve - p.freeze for p in available_parts)
                if available_count >= required_count:
                    for part in available_parts:
                        if required_count <= 0:
                            break
                        if bom.board.status == 'RC':
                            part_cnt = part.count - part.reserve
                        else:
                            part_cnt = part.count - part.reserve - part.freeze
                        use_qty = min(required_count, part_cnt)
                        if use_qty == 0:
                            continue
                        palets_to_create.append(Palet(req=my_req, part=part, quantity=use_qty, bom=bom))
                        part.reserve += use_qty
                        if bom.board.status == 'RC':
                            part.freeze -= use_qty
                        part.changed_request = f"request {my_req.id}"
                        parts_to_update.add(part)
                        required_count -= use_qty
                else:
                    if Pn in rejected_parts:
                        rej.append({
                            'part_number': Pn,
                            'bom_total': required_count,
                            'part_total': rejected_parts[Pn],
                        })
                    insf.append({
                        'part_number': Pn,
                        'bom_total': required_count,
                        'part_total': available_count,
                        'difference': 0,
                    })
            if palets_to_create:
                Palet.objects.bulk_create(palets_to_create)
            Part.objects.bulk_update(list(parts_to_update), ['changed_request','reserve', 'freeze'])
            my_palets = Palet.objects.filter(req=my_req)
            if not my_palets.exists():
                my_req.delete()
                return render(request, 'status.html', {
                    'title': 'No Part Found',
                    'message': 'no part was found for the selected boards'
                })
            Board.objects.filter(id__in=board_ids).update(status="PD")
            distinct_keys = my_palets.values_list('part__order__order_number', 'part__part_number').distinct()
            for i, (order_no, part_no) in enumerate(distinct_keys, start=1):
                plta = PaletAnbar.objects.create(number=i)
                my_palets.filter(part__order__order_number=order_no, part__part_number=part_no).update(palet_anbar=plta)
            def merge_by_part(items):
                merged = {}
                for d in items:
                    pn = d['part_number']
                    if pn in merged:
                        merged[pn]['bom_total'] += d['bom_total']
                    else:
                        merged[pn] = d.copy()
                return list(merged.values())
            insf = merge_by_part(insf)
            rej = merge_by_part(rej)
            if not insf:
                return render(request, 'darkhast_bad.html', {'message': 'All parts are in stock','req': my_req})
            return render(request, 'darkhast_bad.html', {'insufficient_parts': insf,'req': my_req,'rej': rej})

    projects = Project.objects.all()
    boards = Board.objects.filter(status__in=['RB','RC']).select_related("project_name")
    boards_data = dumps([
        {
            "id": board.id,
            "title": board.title,
            "part_number": board.part_number,
            "project_id": board.project_name.id if board.project_name else None
        }
        for board in boards
    ])
    context = {
        'orders': Order.objects.values_list('order_number', flat=True),
        'users': User.objects.filter(groups__name='assemble'),
        'projects': projects,
        'boards_json': boards_data,
    }
    return render(request, 'darkhast.html', context)


@login_required
def darkhast_acc(request):
    has_permision(request, 'assemble')
    action = request.POST.get('action').split('_')
    my_req = get_object_or_404(Darkhast, id=action[1])
    with transaction.atomic():
        if action[0] == 'return':
            Board.objects.filter(bom__palet__req=my_req).distinct().update(status='RB')
            palets = Palet.objects.filter(req=my_req)
            for palet in palets:
                my_part = palet.part
                my_part.reserve -= palet.quantity
                my_part.save()
            palets.delete()
            my_req.delete()
            return redirect('darkhast')
        BOM.objects.filter(palet__req=my_req).update(status='PND')
    return render(request, 'status.html', {'title':'Success','message':f'Request {my_req.pk} was sent to the anbardar'})


@login_required
def darkhast_completion(request):
    has_permision(request, 'admin')
    project = request.GET.get('project')
    title = request.GET.get('title')
    
    page_number = request.GET.get('page', 1)
    
    boards = Board.objects.filter(status__in=['PA','RB','PD'])
    for b in boards:

        b.status = b.is_complete()
    Board.objects.bulk_update(boards, ["status"])
    boards = Board.objects.filter(status='PA')

    if project:
        boards = boards.filter(project_name__name__icontains=project)
    if title:
        boards = boards.filter(title__icontains=title)

    boards = boards.values('project_name__name', 'title').distinct()
    paginator = Paginator(boards, 10)
    page_obj = paginator.get_page(page_number)
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        board_data = [
            {
                'project': b['project_name__name'],
                'title': b['title'],
            }
            for b in page_obj
        ]
        return JsonResponse({
            'boards': board_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })
    return render(request, 'darkhast_completion.html' ,{'boards':page_obj})


@login_required
def darkhast_completion_boards(request, projectname, title):
    has_permision(request, 'admin')
    if request.method == 'POST':

        board_id = request.POST.getlist('selected_boards_hidden')
        selected_order = Order.objects.all().values_list('id', flat=True)
        selected_user = request.POST.get('user')

        boms = BOM.objects.filter(board__in=board_id, status='EPT')
        with transaction.atomic():
            target_user = User.objects.get(id=selected_user)
            my_req = Darkhast.objects.create(user=request.user, for_user=target_user, req_type='COM')
            parts_qs = Part.objects.filter(order__in=selected_order)

            for bom in boms:
                Pn = bom.part_number
                required_count = int(bom.count)
                my_part = parts_qs.filter(part_number=Pn,failure='ACCEPT').order_by('-pk')
            
                for part in my_part:
                    if required_count <= 0:
                        break
                    use_qty = min(required_count, part.count - part.reserve - part.freeze)
                    if use_qty == 0:
                        continue
                    Palet.objects.create(req=my_req, part=part, quantity=use_qty, bom=bom)
                    bom.status = "TV"
                    part.reserve += use_qty 
                    part.changed_request = f"request {my_req.id}"
                    required_count -= use_qty
                Part.objects.bulk_update(my_part, ['reserve','changed_request'])
            BOM.objects.bulk_update(boms, ['status'])
            my_palets = Palet.objects.filter(req=my_req)
            for index, item in enumerate(my_palets.values_list('part__order__order_number', 'part__part_number').distinct()):
                plta=PaletAnbar.objects.create(number=index+1)
                my_palets.filter(part__order__order_number=item[0], part__part_number=item[1]).update(palet_anbar=plta)   
            a = Palet.objects.filter(req=my_req)
        return render(request, 'palet_preview.html', {'palets': a, 'req':my_req})

    boards = Board.objects.filter(project_name__name=projectname, title=title).filter(status='PA').select_related('project_name')
    version = request.GET.get('version')
    serial = request.GET.get('serial')
    page_number = request.GET.get('page', 1)
    if version:
        boards = boards.filter(version__icontains=version)
    if serial:
        boards = boards.filter(serial__icontains=serial)

    boards = boards.annotate(eptboms=Count('bom', filter=Q(bom__status='EPT'))).order_by('eptboms')
    boms = BOM.objects.filter(status='EPT', board__in=boards.values_list("id", flat=True))
    part_count = (Part.objects.filter(failure='ACCEPT',part_number__in=boms.values_list('part_number', flat=True)).values('part_number')
    .annotate(total_available=Sum(F('count') - F('reserve') - F('freeze'))))

    part_count_dict = {p['part_number']: p['total_available'] for p in part_count}
    cmplt = {}
    for b in boms:
        if b.board.part_number in cmplt:
            if b.count < part_count_dict.get(b.part_number, 0):
                cmplt[b.board.part_number] += 1
                part_count_dict[b.part_number] -= b.count
        else:
            cmplt[b.board.part_number] = 0
            if b.count < part_count_dict.get(b.part_number, 0):
                cmplt[b.board.part_number] += 1
                part_count_dict[b.part_number] -= b.count

    paginator = Paginator(boards, 20)
    page_obj = paginator.get_page(page_number)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        board_data = [
            {
                'id': b.id,
                'project': b.project_name.name,
                'title': b.title,
                'part_number': b.part_number,
                'eptboms': b.eptboms,
                'cmplt': cmplt.get(b.part_number, 0)
            }
            for b in page_obj
        ]
        return JsonResponse({
            'boards': board_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })
    return render(request,'darkhast_completion_boards.html',{'boards': page_obj, 'users': User.objects.filter(groups__name='assemble'),})



@login_required
def darkhast_part(request):
    has_permision(request, 'assemble')

    if request.method == 'POST':
        selected_parts = request.POST.getlist('selected_parts')
        errors = []
        part_data = []

        for part_id in selected_parts:
            quantity = request.POST.get(f'quantity_{part_id}')
            if not quantity or not quantity.isdigit():
                return render(request, "status.html", {"title":"Error","message":f"Invalid quantity for part ID {part_id}"})

            quantity = int(quantity)
            my_part = get_object_or_404(Part, pk=part_id)

            available = my_part.count - my_part.reserve - my_part.freeze
            if quantity > available:
                return render(request, "status.html", {"title":"Error","message":f"Not enough stock for part ID {part_id} (available: {available})"})

            part_data.append((my_part, quantity))

        with transaction.atomic():
            my_req = Darkhast.objects.create(user=request.user, req_type='MAF')
            for part, qty in part_data:
                Mafghood.objects.create(req=my_req, quantity=qty, maf_part=part)

        message = 'Request was sent successfully:<br><br>'
        for part, qty in part_data:
            message += f"Code Anbar: {part.pk} | Quantity: {qty}<br>"

        return render(request, 'status.html', {
            'title': 'Success',
            'message': message
        })

    filtered_boms = BOM.objects.select_related('board__project_name').filter(board__status='NR')
    parts = Part.objects.all()#.filter(pk__lt=10101021016) #########################################################################################filter tolidi
    boardname = request.GET.get('boardname')
    project = request.GET.get('project')
    designator = request.GET.get('designator')
    version = request.GET.get('version')
    partnumber = request.GET.get('partnumber')
    codeanbar = request.GET.get('codeanbar')
    page_number = request.GET.get('page', 1)

    if boardname:
        filtered_boms = filtered_boms.filter(board__title__icontains=boardname)
    if project:
        filtered_boms = filtered_boms.filter(board__project_name__name__icontains=project)
    if designator:
        filtered_boms = filtered_boms.filter(designators__icontains=designator)
    if version:
        filtered_boms = filtered_boms.filter(board__version__icontains=version)
    if partnumber:
        filtered_boms = filtered_boms.filter(part_number__icontains=partnumber)
    if codeanbar:
        matching_parts = Part.objects.filter(id__icontains=codeanbar)
        part_numbers = matching_parts.values_list('part_number', flat=True)
        filtered_boms = filtered_boms.filter(part_number__in=part_numbers)
    else:
        matching_parts = parts

    paginator = Paginator(filtered_boms, 10)
    page_obj = paginator.get_page(page_number)

    part_numbers_on_page = page_obj.object_list.values_list('part_number', flat=True)
    parts_on_page = matching_parts.filter(part_number__in=part_numbers_on_page)

    part_map = defaultdict(list)
    for part in parts_on_page:
        part_map[part.part_number].append(part)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        board_data = []
        for bom in page_obj:
            parts = part_map.get(bom.part_number, [])
            if not parts:
                board_data.append({
                    "title": bom.board.title,
                    "project": bom.board.project_name.name,
                    "boardid": bom.board.name,
                    "version": bom.board.version,
                    "designator": bom.designators,
                    "part_number": bom.part_number,
                    "code_anbar": "",
                    "ordernumber": "",
                    "count": "",
                })
            else:
                for part in parts:
                    board_data.append({
                        "title": bom.board.title,
                        "project": bom.board.project_name.name,
                        "boardid": bom.board.name,
                        "version": bom.board.version,
                        "designator": bom.designators,
                        "part_number": bom.part_number,
                        "code_anbar": part.id,
                        "ordernumber": part.order.order_number,
                        "count": part.count - part.reserve,
                    })

        return JsonResponse({
            'boards': board_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })

    return render(request, 'darkhast_part2.html', {
        'boards': page_obj,
    })
