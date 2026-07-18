from django.shortcuts import render, get_object_or_404
from mech.models import Part, PartPack, Operation, PartOperation
from anbar.models import Board, Project, Darkhast, BOM, SepehrQC
from django.db.models.functions import Concat
from openpyxl import Workbook
from json import dumps
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.db.models import Count, F, Value, OuterRef, Subquery
from collections import defaultdict
from django.core.paginator import Paginator
from django.http import JsonResponse
from .models import *
from django.core.exceptions import ValidationError
from pandas import read_excel
from .permisions import has_permision
import re
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

@login_required
def elec_parts(request):
    boards = Board.objects.filter(status='AS')
    return render(request, 'assembly.html', {'boards': boards})


@login_required
def mech_parts(request):
    parts = Part.objects.filter(status='ASM')
    return render(request, 'mech_parts_assembly.html', {'parts': parts})


@login_required
def add_tree(request):
    has_permision(request, 'admin')
    projects = Project.objects.all()
    if request.method == 'POST':
        new_project = request.POST.get('new_project')
        old_project = request.POST.get('old_project')
        uploaded_file = request.FILES.get('file')
        try:
            with transaction.atomic():
                df = read_excel(uploaded_file)
                df = df.fillna('#')
                if new_project:
                    if Project.objects.filter(name=new_project).exists():
                        return render(request, 'add_tree.html', {"error":f'{new_project} Project already exist', 'projects':projects})
                    project = Project.objects.create(name=new_project)
                elif old_project:
                    project = Project.objects.get(name=old_project)
                else:
                    return render(request, 'add_tree.html', {"error":'please choose a project', 'projects':projects})
                Tree.objects.filter(project=project).delete()

                electronic = zip(df['Electronic'][1:], df['Unnamed: 1'][1:])
                mechanic = zip(df['Mechanic'][1:], df['Unnamed: 3'][1:])
                software = df['Software'][1:]
                all_tree = []

                elec_pattern = r"E\d{3}$"
                for board, qnt in electronic:
                    if board != '#' and qnt != '#':
                        if re.match(elec_pattern, board) and isinstance(qnt, int):
                            all_tree.append(Tree(project=project, department='ELC', partnumber=board, quantity=qnt))
                        else:
                            raise ValidationError("board ID or quantity did not match the pattern")
                    else:
                        continue
                mech_pattern = r"M\d{3}$"
                for part, qnt in mechanic:
                    if part != '#' and qnt != '#':
                        if re.match(mech_pattern, part) and isinstance(qnt, int):
                            all_tree.append(Tree(project=project, department='MEC', partnumber=part, quantity=qnt))
                        else:
                            raise ValidationError("part ID or quantity did not match the pattern")
                    else:
                        continue
                soft_pattern = r"B\d{3}$"
                for soft in software:
                    if soft != '#':
                        if re.match(soft_pattern, soft):
                            all_tree.append(Tree(project=project, department='SFT', partnumber=soft, quantity=1))
                        else:
                            raise ValidationError("software ID did not match the pattern")
                    else:
                        continue
                Tree.objects.bulk_create(all_tree)
                elec_tree = [i for i in all_tree if i.department=='ELC']
                mech_tree = [i for i in all_tree if i.department=='MEC']
                soft_tree = [i for i in all_tree if i.department=='SFT']
        except Exception as e:
            return render(request, 'add_tree.html', {"error":str(e), 'projects':projects})
        return render(request, 'tree_postview.html', {'elec_tree':elec_tree, "mech_tree":mech_tree, "soft_tree":soft_tree})
    #GET
    return render(request, 'add_tree.html', {'projects': projects})


@login_required
def add_template(request): 
    wb = Workbook()
    ws = wb.active
    ws.title = "Tree"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Electronic"
    ws.merge_cells("C1:D1")
    ws["C1"] = "Mechanic"
    ws["E1"] = "Software"
    ws["A2"] = "PartNumber"
    ws["B2"] = "Quantity"
    ws["C2"] = "PartNumber"
    ws["D2"] = "Quantity"
    ws["E2"] = "ID"
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 15
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=tree_template.xlsx'
    wb.save(response)
    return response


@login_required
def all_tree(request):
    projects = Project.objects.all()
    if request.method == 'POST':
        proj = request.POST.get('project')
        tree = Tree.objects.filter(project__name=proj)
        elec_tree = [i for i in tree if i.department=='ELC']
        mech_tree = [i for i in tree if i.department=='MEC']
        soft_tree = [i for i in tree if i.department=='SFT']
        return render(request, 'all_tree.html', {"projects":projects, 'elec_tree':elec_tree, "mech_tree":mech_tree, "soft_tree":soft_tree, 'thisproject':proj})
    #GET
    return render(request, 'all_tree.html', {"projects":projects})


@login_required
def create_product(request):
    projects = Project.objects.all()
    if request.method == 'POST':
        proj = request.POST.get('project')
        tree = Tree.objects.filter(project__name=proj, department='ELC')
        mech_products = Product.objects.filter(project__name=proj, boardpack__isnull=True)
        for t in tree:
            tree_id = t.partnumber[1:]
            related_parts = Board.objects.filter(name=tree_id, status='AS').values_list('part_number', flat=True)
            t.rel_parts = related_parts
        return render(request, 'create_product.html', {"projects":projects, 'tree':tree, 'thisproject':proj, 'mech_products':mech_products})
    #GET
    return render(request, 'create_product.html', {"projects":projects})


@login_required
def edit_product(request, pk):
    my_product = get_object_or_404(Product, id=pk)
    my_project = my_product.project.name
    tree = Tree.objects.filter(project__name=my_project, department='ELC')
    mech_products = Product.objects.filter(project__name=my_project, boardpack__isnull=True)
    current_boards = Board.objects.filter(boardpack__product=my_product)
    for t in tree:
        tree_id = t.partnumber[1:]
        related_parts = Board.objects.filter(name=tree_id, status='AS').values_list('part_number', flat=True)
        current_parts = current_boards.filter(name=tree_id).values_list('part_number', flat=True)
        t.rel_parts = related_parts
        t.current_parts = current_parts
    return render(request, 'edit_product.html', {'tree':tree, 'thisproject':my_project, 'mech_products':mech_products})


@login_required
def save_product_assembly(request):
    board_list = request.POST.get('data')
    product = request.POST.get('product')
    project = request.POST.get('project')
    proname = request.POST.get('product_name')
    board_qs = Board.objects.filter(part_number__in=board_list.split(','))
    my_project = Project.objects.get(name=project)

    my_product = Product.objects.get(project=my_project, productnumber=product)
    all_packs = []
    for board in board_qs:
        board.status = 'FN'
        pack = BoardPack(board=board, product=my_product)
        all_packs.append(pack)
    BoardPack.objects.bulk_create(all_packs)
    Board.objects.bulk_update(board_qs, ['status'])
    mech_packs = PartPack.objects.filter(product=my_product)
    my_product.productnumber = proname
    my_product.save()
    return render(request, 'product_postview_as.html', {"product":my_product, "board_packs":all_packs, "mech_packs":mech_packs})


@login_required
def all_products(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        Board.objects.filter(boardpack__product__id=product_id).update(status='AS')
        Product.objects.get(id=product_id).delete()

    products = Product.objects.annotate(boardpack_count=Count('boardpack')).filter(boardpack_count__gt=0)
    return render(request, 'all_products_as.html', {"products":products})


@login_required
def product_details(request, pk):
    my_product = get_object_or_404(Product, id=pk)
    part_packs = PartPack.objects.filter(product=my_product)
    board_packs = BoardPack.objects.filter(product=my_product)

    board_dict = defaultdict(list)
    for board in board_packs:
        board_dict[board.board.title].append(board.board.part_number)
    
    part_dict = defaultdict(list)
    for part in part_packs:
        part_dict[part.part.title].append(part.part.part_number)
    
    part_dict = dict(part_dict)
    board_dict = dict(board_dict)

    return render(request, 'product_details_as.html', {"part_packs":part_dict, "board_packs":board_dict, "product":my_product})



@login_required
def board_product_details(request, PN):
    my_board = get_object_or_404(Board, part_number=PN)
    my_board.requests = Darkhast.objects.filter(req_type__in=['ASS','MAF','COM']).filter(palet__bom__board=my_board).distinct()
    return render(request, 'board_product_details.html', {"board":my_board})


@login_required
def part_product_details(request, PN):
    my_part = get_object_or_404(Part, part_number=PN)
    part_op = PartOperation.objects.filter(part=my_part)
    sheet_op = Operation.objects.filter(material__part=my_part)
    return render(request, 'part_product_details.html', {"board":my_part, "part_op":part_op, "sheet_op":sheet_op})


@login_required
def bom_details(request, pk):
    my_board = get_object_or_404(Board, id=pk)
    boms = BOM.objects.filter(board=my_board)

    designator = request.GET.get('designator')
    part_number = request.GET.get('part_number')
    description = request.GET.get('description')
    status = request.GET.get('status')
    page_number = request.GET.get('page', 1)

    if designator:
        boms = boms.filter(designators__icontains=designator)
    if part_number:
        boms = boms.filter(part_number__icontains=part_number)
    if description:
        boms = boms.filter(description__icontains=description)
    if status:
        boms = boms.filter(status__in=status.split(','))

    paginator = Paginator(boms, 20)
    page_obj = paginator.get_page(page_number)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        bom_data = [
            {   
                'part_number': p.part_number,
                'designators' : p.designators,
                'description' : p.description,
                'status' : p.get_status_display(),
            }
            for p in page_obj
        ]
        return JsonResponse({
            'boms': bom_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'page_size' : 20 ,
            'current_page': page_obj.number,
        })
    
    projectname = my_board.project_name.name
    return render(request, 'bom_details.html', {
        'boms': page_obj,
        'project':f"{projectname} - {my_board.title}",
        'title':my_board.part_number,
    })


@login_required
def product_flow(request):
    all_projects = Project.objects.values_list('name', flat=True)
    if request.method == 'POST':
        project = request.POST.get('projectname')
        parts = Part.objects.filter(project_name__name=project)
        boards = Board.objects.filter(project_name__name=project)
        grouped_parts = {stat:list(parts.filter(status=stat).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True)) for stat in Part.objects.values_list('status', flat=True)}
        grouped_parts['HCK'] = list(parts.annotate(last_description=Subquery(PartOperation.objects.filter(part=OuterRef('pk')).order_by('-id').values('description')[:1])).filter(last_description__icontains="حک لیزر").annotate(part_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('part_str', flat=True))
        grouped_parts['AAB'] = list(parts.annotate(last_description=Subquery(PartOperation.objects.filter(part=OuterRef('pk')).order_by('-id').values('description')[:1])).filter(last_description__icontains="آبکاری").annotate(part_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('part_str', flat=True))
        grouped_boards = {stat:list(boards.filter(status=stat).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True)) for stat in ['OR','RB','PD','AS','FN']}
        grouped_boards['PA'] = list(boards.annotate(num_operations=Count('boardoperation')).filter(status='PA', num_operations=0).exclude(bom__palet__req__req_type='COM').annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['CA'] = list(boards.annotate(num_operations=Count('boardoperation')).filter(status='CA', num_operations=0).exclude(bom__palet__req__req_type='COM').annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['TK'] = list(boards.annotate(num_operations=Count('boardoperation')).filter(bom__palet__req__req_type='COM', num_operations=0).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['SS'] = list(boards.annotate(num_operations=Count('boardoperation')).filter(num_operations__gt=0).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['QC'] = list(boards.filter(status='QC', qc_req__isnull=True).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['OP'] = list(boards.filter(status='QC', qc_req__isnull=False).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['SPS'] = list(boards.filter(pk__in=Subquery(SepehrQC.objects.filter(board=OuterRef('pk')).order_by('-date').filter(status='NOR').values('board'))).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        grouped_boards['SPR'] = list(boards.filter(pk__in=Subquery(SepehrQC.objects.filter(board=OuterRef('pk')).order_by('-date').filter(status__in=['ACC','PND','REJ']).values('board'))).annotate(board_str=Concat(F('title'), Value('-V'), F('version'), Value('S'), F('serial'), Value('\n'))).values_list('board_str', flat=True))
        products = list(Product.objects.filter(project__name=project).annotate(product_str=Concat(F('project__name'), Value('-'), F('productnumber'), Value('\n'))).values_list('product_str', flat=True))
    
        return render(request, 'product_flow.html', {'selected_project':project,'projects':all_projects, "boardstages":dumps(dict(grouped_boards)), "partstages":dumps(dict(grouped_parts)), "products":dumps(products)})
    return render(request, 'product_flow.html', {'projects':all_projects})


@login_required
def flow_details(request, project, filter):
    type, fil = filter.split('_')
    if type == 'board':
        boards = Board.objects.filter(project_name__name=project)
        if fil == 'pa':
            data = boards.annotate(num_operations=Count('boardoperation')).filter(status='PA', num_operations=0).exclude(bom__palet__req__req_type='COM')
        elif fil == 'ca':
            data = boards.annotate(num_operations=Count('boardoperation')).filter(status='CA', num_operations=0).exclude(bom__palet__req__req_type='COM')
        elif fil == 'tk':
            data = boards.annotate(num_operations=Count('boardoperation')).filter(bom__palet__req__req_type='COM', num_operations=0)
        elif fil == 'ss':
            data = boards.annotate(num_operations=Count('boardoperation')).filter(num_operations__gt=0)
        elif fil == 'qc':
            data = boards.filter(status='QC', qc_req__isnull=True)
        elif fil == 'op':
            data = boards.filter(status='QC', qc_req__isnull=False)
        elif fil == 'sps':
            data = boards.filter(pk__in=Subquery(SepehrQC.objects.filter(board=OuterRef('pk')).order_by('-date').filter(status='NOR').values('board')))
        elif fil == 'spr':
            data = boards.filter(pk__in=Subquery(SepehrQC.objects.filter(board=OuterRef('pk')).order_by('-date').filter(status__in=['ACC','PND','REJ']).values('board')))
        else:
            data = boards.filter(status=fil.upper())
    else:
        parts = Part.objects.filter(project_name__name=project)
        if fil == 'hck':
            data = parts.annotate(last_description=Subquery(PartOperation.objects.filter(part=OuterRef('pk')).order_by('-id').values('description')[:1])).filter(last_description__icontains="حک لیزر")
        elif fil == 'aab':
            data = parts.annotate(last_description=Subquery(PartOperation.objects.filter(part=OuterRef('pk')).order_by('-id').values('description')[:1])).filter(last_description__icontains="آبکاری")
        else:
            data = parts.filter(status=fil.upper())
    return render(request, 'flow_details.html', {"data":data})