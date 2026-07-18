from django.shortcuts import render, redirect
from anbar.permisions import has_permision
from anbar.models import Part, Board, BOM, Project, Order
from .models import *
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
from django.http import HttpResponseBadRequest, HttpResponseNotAllowed
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum, Subquery, OuterRef, F, Case, When, Value, IntegerField
from pandas import read_excel
from django.db import transaction
from json import dumps, loads
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt


@login_required
def finance_estimate(request):
    has_permision(request, 'finance')
    if request.method == 'POST':

        board_id = request.POST.get('board')
        orders = request.POST.getlist('order')
        quantity = int(request.POST.get('quantity'))
        my_board = Board.objects.get(id=board_id)
        all_boms = BOM.objects.filter(board=my_board)
        bom_summary = (all_boms.values('part_number').annotate(bom_total=Sum('count')))
        orders_name = list(Order.objects.filter(id__in=orders).values_list('order_number', flat=True))
        part_summary = {
            p['part_number']: p['part_total']
            for p in Part.objects.filter(order__in=orders).exclude(failure='REJ').values('part_number').annotate(part_total=Sum('count')-Sum('reserve')-Sum('freeze'))}
        insf = []
        with transaction.atomic():
            
            my_req = DarkhastFinance.objects.create(board = my_board, sets=quantity)
            for bom in bom_summary:
                part_number = bom['part_number']
                bom_total = quantity * bom['bom_total'] or 0
                part_total = part_summary.get(part_number, 0)
                difference = part_total - bom_total  #1.2*bom_total
                if difference < 0:
                    my_palet = PaletFinance(part_number=part_number, quantity=abs(difference), req=my_req, description='-') #quantity=int(math.ceil(abs(difference*1.2)))
                    insf.append(my_palet)

                part_qs = Part.objects.filter(part_number=part_number, order__in=orders)
                for part in part_qs:
                    if bom_total <= 0:
                        break
                    used, _ = PaletPartFinance.objects.get_or_create(part=part, req=my_req)
                    min_part = min(part.count - part.reserve - part.freeze - used.quantity, bom_total)
                    if min_part == 0:
                        continue
                    used.quantity += min_part
                    used.save()
                    bom_total -= min_part
            PaletFinance.objects.bulk_create(insf)
            new_bom = []
            for _ in range(quantity):
                my_board = Board.objects.create(project_name=my_board.project_name, title=my_board.title, name=my_board.name, version=my_board.version, status='OR')
                UsedOrders.objects.create(board=my_board, orders=orders_name)
                for b in all_boms:
                    b_dict = model_to_dict(b)
                    b_dict.pop('id', None)
                    b_dict['board'] = my_board
                    new_bom.append(BOM(**b_dict))
            BOM.objects.bulk_create(new_bom)
        return render(request, 'finance_deflicts.html', {'palets':insf, 'req':my_req})

    projects = Project.objects.all()
    boards = Board.objects.filter(status='NR').select_related("project_name")
    boards_data = dumps([
        {
            "id": board.id,
            "title": board.title,
            "version": board.version,
            "project_id": board.project_name.id if board.project_name else None}
        for board in boards])
    context = {
        'orders': Order.objects.all(),
        'projects': projects,
        'boards_json': boards_data,}
    
    return render(request, 'finance_estimate.html', context)


@login_required
def finance_acc(request):
    has_permision(request, 'finance')
    rej = request.POST.get('rej')
    acc = request.POST.get('acc')

    if acc:
        my_req = DarkhastFinance.objects.get(id=acc)
        my_fin = PaletPartFinance.objects.filter(req=my_req)
        for fin in my_fin:
            my_part = fin.part
            my_part.freeze += fin.quantity
            my_part.save()
        my_fin.delete()
        return render(request, 'status.html', {'title':'Success','message':f'Request {acc} was sent to the finance'})
    else:
        DarkhastFinance.objects.filter(id=rej).delete()
        return redirect("finance_estimate")
    

@login_required
def panel_finance(request):
    has_permision(request, 'finance')
    if request.method == 'POST':
        req_id = request.POST.get('darkhast')
        my_req = get_object_or_404(DarkhastFinance, id=req_id)
        my_palets = PaletFinance.objects.filter(req=my_req).distinct()
        wb = Workbook()
        ws = wb.active
        ws.title = str(my_req.id)+ 'finance request'
        headers = ['partnumber', 'quantity']
        ws.append(headers)
        for palet in my_palets:
            ws.append([palet.part_number, palet.quantity])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={my_req.id}_finance_request.xlsx'
        wb.save(response)
        return response

    darkhasts = DarkhastFinance.objects.all()
    return render(request, 'darkhast_list_finance.html', {'darkhasts':darkhasts})


@login_required
def panel_finance_details(request, pk):
    has_permision(request, 'finance')
    my_req = get_object_or_404(DarkhastFinance, id=pk)
    palets = PaletFinance.objects.filter(req=my_req)
    return render(request, 'palet_finance.html', {"palets":palets, "my_req":my_req})


@login_required
def panel_finance_acc(request, pk):
    has_permision(request, 'finance')
    my_req = get_object_or_404(DarkhastFinance,id=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        palets = request.POST.getlist('palets')
        order = request.POST.get('order')
        new_order = request.POST.get('new')

        if action == 'button1':
            if order == None and not new_order:
                    return render(request, 'status.html', {'title':'Error','message':'please enter a new order or choose one of existing orders'})
            with transaction.atomic():

                my_req.status = "ACC"
                my_req.save()
                if order == None:
                    my_order = Order.objects.create(order_number = new_order)
                    for ID in palets:    
                        my_plt = PaletFinance.objects.get(id=int(ID))
                        Part.objects.create(order=my_order, part_number=my_plt.part_number, ordered=my_plt.quantity, count=0)
                else:
                    my_order = Order.objects.get(id=order)
                    palet_qs = PaletFinance.objects.filter(id__in=list(map(int,palets)))
                    for my_plt in palet_qs:
                        try:
                            my_part = Part.objects.get(order=my_order, part_number=my_plt.part_number)
                        except Part.MultipleObjectsReturned:
                            return render(request, 'status.html', {'title':'Error','message':'multiple Part Object with same PN and ON was found'})
                        except Part.DoesNotExist:
                            Part.objects.create(order=my_order, part_number=my_plt.part_number, ordered=my_plt.quantity, count=0)
                            continue
                        ord = my_part.ordered + my_plt.quantity
                        my_part.update(ordered=ord)
            return render(request, 'status.html', {'title':'Success','message':'Request was Accepted Successfuly'})
        
        elif action == 'button2':
            my_req.status = "REJ"
            my_req.save()
            return render(request, 'status.html', {'title':'Success','message':'Request was Rejected Successfuly'})
        return HttpResponseBadRequest("Bad request: invalid input.")
    return HttpResponseNotAllowed(['POST'])
    
@csrf_exempt
def order_list(request):
    has_permision(request, 'finance')
    orders = Order.objects.all()
    if request.method == 'POST':
        data = loads(request.body)
        orderid = data['orderId']
        my_order = Order.objects.get(id=orderid)
        my_order.vendor = data['vendor'] or None
        my_order.status = data['status']
        my_order.track_number = data['track_number'] or None
        my_order.peyment_date = data['peyment_date'] or None
        my_order.arrive_date = data['arrive_date'] or None
        my_order.supply_date = data['supply_date'] or None
        my_order.resupply_date = data['resupply_date'] or None
        my_order.save()
        return render(request, 'order_list.html', {'orders':orders})
    
    vendor = request.GET.get('vendor')
    track_number = request.GET.get('track_number')
    order_number = request.GET.get('order_number')
    status = request.GET.get('failure')
    page_number = request.GET.get('page', 1)

    if vendor:
        orders = orders.filter(vendor__icontains=vendor)
    if track_number:
        orders = orders.filter(track_number__icontains=track_number)
    if order_number:
        orders = orders.filter(order_number__icontains=order_number)  
    if status:
        orders = orders.filter(status__in=status.split(','))

    paginator = Paginator(orders, 15)
    page_obj = paginator.get_page(page_number)
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        part_data = [
            {   
                "pk":p.pk,
                'order_number':p.order_number,
                'status' : p.get_status_display(),
                'vendor' : p.vendor or '-',
                'track_number' : p.track_number or '-',
                'order_date':p.order_date or '-',
                'peyment_date':p.peyment_date or '-',
                'arrive_date':p.arrive_date or '-',
                'supply_date':p.supply_date or '-',
                'resupply_date':p.resupply_date or '-',
            }
            for p in page_obj
        ]
        return JsonResponse({
            'parts': part_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })
    
    return render(request, "order_list.html", {"orders":page_obj})


@login_required
def order_del(request, pk):
    has_permision(request, 'finance')
    my_order = Order.objects.get(id=pk)
    if my_order.status == 'RCV':
        return redirect('order_list')
    my_order.delete()
    return redirect('order_list') 


@login_required
def order_details(request, pk):
    has_permision(request, 'finance')
    my_order = get_object_or_404(Order, id=pk)
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        try:
            with transaction.atomic():
                df = read_excel(uploaded_file)
                PaymentReport.objects.filter(order_id=pk).delete()
                for _, part in df.iterrows():
                    part_number=part['Part Number']
                    quantity=part['Quantity']
                    price = part['Price']
                    PaymentReport.objects.create(part_number=part_number, quantity=quantity, price=price, order=my_order)
        except Exception as e:
            return render(request, 'finance_status.html', {"message":str(e), "title":"Error"})
        return redirect('payment_details', pk=pk)
    #GET
    my_palets = PaletFinance.objects.filter(order=my_order, not_delivered=False)
    palet_data = list(my_palets.values('part_number', 'quantity', 'status', 'description', 'req__board__project_name__name', 'req__board__title', 'req__board__version'))
    compact = my_palets.values('part_number').distinct().annotate(total=Sum('quantity'))
    return render(request, 'order_details.html', {'palets':dumps(palet_data), 'compact':compact, 'order':my_order})


@login_required
def export_order(request, pk):
    has_permision(request, 'finance')
    my_order = get_object_or_404(Order, id=pk)
    my_palets = PaletFinance.objects.filter(order=my_order, not_delivered=False).values('part_number').distinct().annotate(total=Sum('quantity'))
    wb = Workbook()
    ws = wb.active
    ws.title = str(my_order.order_number)
    headers = ['partnumber', 'quantity']
    ws.append(headers)
    for palet in my_palets:
        ws.append([palet['part_number'], palet['total']])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={my_order.order_number}.xlsx'
    wb.save(response)
    return response
    

@login_required
def templates(request):
    has_permision(request, 'finance')
    if request.method == 'POST':
        temp_type = request.POST.get('type')
        if temp_type == "payment":
            wb = Workbook()
            ws = wb.active
            ws.title = "Payment Template"
            ws.append(["Part Number", "Quantity", "Price"])
            for col in range(1, 3):
                ws.column_dimensions[get_column_letter(col)].width = 15
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
            response['Content-Disposition'] = 'attachment; filename=Payment_Template.xlsx'
            wb.save(response)
            return response
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Shipment Template"
            ws.append(["Part Number", "Quantity"])
            for col in range(1, 2):
                ws.column_dimensions[get_column_letter(col)].width = 15
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
            response['Content-Disposition'] = 'attachment; filename=Shipment_Template.xlsx'
            wb.save(response)
            return response


@login_required
def payment_details(request, pk):
    has_permision(request, 'finance')
    my_order = get_object_or_404(Order, id=pk)
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        payment_data = loads(request.POST.get('payment_in_data'))
        try:
            with transaction.atomic():
                df = read_excel(uploaded_file)
                for item in payment_data:
                    if item.get('shortage', 0) > 0:
                        PaletFinance.objects.create(part_number=item['part_number'], quantity=item['shortage'], order_id=f"{my_order.order_number}-shortage", not_delivered=True)
                ShipmentReport.objects.filter(order_id=pk).delete()
                for _, part in df.iterrows():
                    part_number=part['Part Number']
                    quantity=part['Quantity']
                    ShipmentReport.objects.create(part_number=part_number, quantity=quantity, order=my_order)
        except Exception as e:
            return render(request, 'finance_status.html', {"message":str(e), "title":"Error"})
        return redirect('shipment_details', pk=pk)
    
    palet_qs = PaletFinance.objects.filter(order=my_order, not_delivered=False).values('part_number').distinct().annotate(total=Sum('quantity'))
    payment_qs = PaymentReport.objects.filter(order=my_order)
    payment_in = payment_qs.annotate(palet_quantity=Subquery(palet_qs.filter(part_number=OuterRef('part_number')).values('total')[:1])).annotate(shortage=Case(When(palet_quantity__gt=F('quantity'), then=F('palet_quantity') - F('quantity')),default=Value(0),output_field=IntegerField()))
    payment_out = PaymentReport.objects.filter(order=my_order).exclude(part_number__in=palet_qs.values_list('part_number', flat=True).distinct())
    palet_out = palet_qs.exclude(part_number__in=payment_qs.values_list('part_number', flat=True).distinct())

    return render(request, 'payment_details.html', {'palet_out':palet_out, 'payment_in':payment_in, 'payment_in_json':dumps(list(payment_in.values())), 'payment_out':payment_out, 'order':my_order})


@login_required
def shipment_details(request, pk):
    has_permision(request, 'finance')
    my_order = get_object_or_404(Order, id=pk)
    if request.method == 'POST':
        shipped = ShipmentReport.objects.filter(order=my_order).annotate(palet_quantity=Subquery(payment_qs.filter(part_number=OuterRef('part_number')).values('total')[:1]))
        for part in shipped:
            Part.objects.create(part_number=part.part_number, count=part.total, order=my_order)
        my_order.status = 'RCV'
        my_order.save()

    payment_qs = PaymentReport.objects.filter(order=my_order).values('part_number').distinct().annotate(total=Sum('quantity'))
    ship_qs = ShipmentReport.objects.filter(order=my_order)
    shipment_in = ship_qs.annotate(palet_quantity=Subquery(payment_qs.filter(part_number=OuterRef('part_number')).values('total')[:1])).annotate(shortage=Case(When(palet_quantity__gt=F('quantity'), then=F('palet_quantity') - F('quantity')),default=Value(0),output_field=IntegerField()))
    shipment_out = ShipmentReport.objects.filter(order=my_order).exclude(part_number__in=payment_qs.values_list('part_number', flat=True).distinct())
    payment_out = PaymentReport.objects.filter(order=my_order).exclude(part_number__in=ship_qs.values_list('part_number', flat=True).distinct())
    return render(request, 'shipment_part.html', {'shipment_in':shipment_in, 'shipment_out':shipment_out, 'payment_out':payment_out, 'order':my_order})


@login_required
def board_order_list(request):
    has_permision(request, 'finance')
    if request.method == 'POST':
        boards_id = request.POST.getlist('req_boards')
        selected_boards = Board.objects.filter(id__in=boards_id)
        with transaction.atomic():
            for b in selected_boards:
                serial = request.POST.get(f"serial_{b.id}")
                b.serial = serial
                b.part_number = f"E{b.name}V{b.version}S{serial}"
                b.status = "RC"
                b.save()
        return render(request, "status.html", {"title":"Success", "message":f"Serials {[b.serial for b in selected_boards]} was Received"})
    boards = Board.objects.filter(status='OR')
    for b in boards:
        used_order = UsedOrders.objects.get(board=b)
        if len(used_order.orders.split(',')) == Order.objects.count():
            b.orders = 'All'
        else:
            b.orders = used_order.orders
    boards_title = boards.values('project_name__name' ,'title', 'version').distinct()
    return render(request, 'board_order_list.html', {'boards':boards, 'titles':boards_title})


@login_required
def draft(request):   
    if request.method == 'POST':
        palet_ids = request.POST.get("paletids")
        order_name = request.POST.get("ordername")
        vendor = request.POST.get("vendor")
        my_order = Order.objects.create(order_number=order_name, vendor=vendor)
        PaletFinance.objects.filter(id__in=palet_ids.split(',')).update(order=my_order)
        return render(request, "finance_status.html", {'title':'success', 'message':f'Order "{order_name}" was created.'})
    
    palets = PaletFinance.objects.filter(order__isnull=True).order_by('part_number') 
    board = request.GET.get('board')
    part_number = request.GET.get('part_number')
    project = request.GET.get('project')
    status = request.GET.get('failure')
    page_number = request.GET.get('page', 1)

    if board:
        palets = palets.filter(req__board__title__icontains=board)
    if part_number:
        palets = palets.filter(part_number__icontains=part_number)
    if project:
        palets = palets.filter(req__board__project_name__name__icontains=project)
    if status:
        palets = palets.filter(status__in=status.split(','))

    paginator = Paginator(palets, 20)
    page_obj = paginator.get_page(page_number)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        part_data = [
            {   
                'pk': p.pk,
                'part_number': p.part_number, 
                'status': p.get_status_display(),
                'quantity': p.quantity,
                'board': f"{p.req.board.title}-V{p.req.board.version}" if p.req else '',
                'project': p.req.board.project_name.name if p.req and p.req.board and p.req.board.project_name else '',
                'sets': p.req.sets if p.req else 0,
                'description': p.description if p.description else '',
            }
            for p in page_obj
        ]
        return JsonResponse({
            'parts': part_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })
    
    return render(request, "draft.html", {"palets":page_obj})


@csrf_exempt
def update_part(request, pk):
    if request.method == 'POST':
        data = loads(request.body)
        quantity = data.get('quantity')
        status = data.get('status')
        description = data.get('description')

        if quantity is None or status is None:
            return HttpResponseBadRequest("Missing data")

        palet = PaletFinance.objects.get(id=pk)
        palet.description = description
        palet.quantity = quantity
        palet.status = status
        palet.save()

        return JsonResponse({
            'quantity': palet.quantity,
            'status': palet.status,
            'status_display': palet.get_status_display()
        })

    return JsonResponse({'error': 'Invalid method'}, status=405)